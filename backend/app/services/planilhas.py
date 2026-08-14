import csv
import json
from datetime import date, timedelta
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from app.models.cartao_ponto import CartaoPonto
from app.models.holerite import Holerite


COR_CABECALHO = "173772"
COR_AVISO = "FFF3CD"
COR_SEQUENCIA = "F8D7DA"
COR_BORDA_SEQUENCIA = "DC3545"


def contem_incerteza(value: object) -> bool:
    return "?" in json.dumps(
        value,
        ensure_ascii=False,
    )


def interpretar_data(value: str) -> date | int | None:
    if "?" in value:
        return None

    if value.isdigit():
        return int(value)

    for separador in ("/", "-"):
        partes = value.split(separador)

        if len(partes) == 3 and all(
            parte.isdigit() for parte in partes
        ):
            dia, mes, ano = map(int, partes)

            try:
                return date(ano, mes, dia)
            except ValueError:
                return None

    return None


def avisos_cartao(cartao: CartaoPonto) -> list[str | None]:
    avisos = []

    for page in cartao.pages:
        anterior: date | int | None = None

        for day in page.days:
            atual = interpretar_data(day.date_raw)
            sequencia_invalida = False

            if atual is not None and anterior is not None:
                esperado = (
                    anterior + timedelta(days=1)
                    if isinstance(anterior, date)
                    else anterior + 1
                )
                sequencia_invalida = atual != esperado

            if atual is not None:
                anterior = atual

            if sequencia_invalida:
                avisos.append("sequencia")
            elif len(day.punches) % 2 != 0 or contem_incerteza(
                day.model_dump()
            ):
                avisos.append("aviso")
            else:
                avisos.append(None)

    return avisos


def avisos_holerite(holerite: Holerite) -> list[str | None]:
    avisos = []
    anterior: tuple[int, int] | None = None

    for page in holerite.pages:
        atual = None

        if page.month.isdigit() and page.year.isdigit():
            mes = int(page.month)
            ano = int(page.year)

            if 1 <= mes <= 12:
                atual = (ano, mes)

        sequencia_invalida = False

        if atual is not None and anterior is not None:
            ano_anterior, mes_anterior = anterior
            esperado = (
                (ano_anterior + 1, 1)
                if mes_anterior == 12
                else (ano_anterior, mes_anterior + 1)
            )
            sequencia_invalida = atual != esperado

        if atual is not None:
            anterior = atual

        if sequencia_invalida:
            avisos.append("sequencia")
        elif (
            not page.fields
            and not page.bases
        ) or contem_incerteza(page.model_dump()):
            avisos.append("aviso")
        else:
            avisos.append(None)

    return avisos


def linhas_cartao(
    cartao: CartaoPonto,
) -> tuple[list[str], list[list[str]]]:
    maior_quantidade = max(
        (
            len(day.punches)
            for page in cartao.pages
            for day in page.days
        ),
        default=0,
    )
    cabecalho = ["Data"]

    for indice in range(maior_quantidade):
        numero = indice // 2 + 1
        tipo = "Entrada" if indice % 2 == 0 else "Saída"
        cabecalho.append(f"{tipo} {numero}")

    linhas = []

    for page in cartao.pages:
        for day in page.days:
            horarios = [
                punch.time_hhmm
                for punch in day.punches
            ]
            linhas.append(
                [day.date_raw]
                + horarios
                + [""] * (maior_quantidade - len(horarios))
            )

    return cabecalho, linhas


def linhas_holerite(
    holerite: Holerite,
) -> tuple[list[str], list[list[str]]]:
    labels = []

    for page in holerite.pages:
        for field in page.fields:
            if field.label not in labels:
                labels.append(field.label)

    cabecalho = ["Pág.", "Mês", "Ano", *labels]
    linhas = []

    for page in holerite.pages:
        valores = {
            field.label: field.value
            for field in page.fields
        }
        linhas.append(
            [page.page, page.month, page.year]
            + [valores.get(label, "") for label in labels]
        )

    return cabecalho, linhas


def gerar_tabela(
    tipo: str,
    value: dict,
) -> tuple[list[str], list[list[str]]]:
    if tipo == "cartao-ponto":
        return linhas_cartao(
            CartaoPonto.model_validate(value)
        )

    if tipo == "holerite":
        return linhas_holerite(
            Holerite.model_validate(value)
        )

    raise ValueError("Tipo de transcrição não suportado.")


def gerar_csv(tipo: str, value: dict) -> bytes:
    cabecalho, linhas = gerar_tabela(tipo, value)
    arquivo = StringIO(newline="")
    escritor = csv.writer(arquivo)
    escritor.writerow(cabecalho)
    escritor.writerows(linhas)

    return arquivo.getvalue().encode("utf-8-sig")


def gerar_xlsx(tipo: str, value: dict) -> bytes:
    cabecalho, linhas = gerar_tabela(tipo, value)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Transcrição"
    worksheet.append(cabecalho)

    for linha in linhas:
        worksheet.append(linha)

    preenchimento = PatternFill(
        fill_type="solid",
        fgColor=COR_CABECALHO,
    )

    for cell in worksheet[1]:
        cell.fill = preenchimento
        cell.font = Font(color="FFFFFF", bold=True)

    if tipo == "cartao-ponto":
        avisos = avisos_cartao(
            CartaoPonto.model_validate(value)
        )
    else:
        avisos = avisos_holerite(
            Holerite.model_validate(value)
        )

    for numero_linha, aviso in enumerate(
        avisos,
        start=2,
    ):
        if aviso is None:
            continue

        cor = (
            COR_SEQUENCIA
            if aviso == "sequencia"
            else COR_AVISO
        )

        for cell in worksheet[numero_linha]:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=cor,
            )

        if aviso == "sequencia":
            worksheet.cell(numero_linha, 1).border = Border(
                left=Side(
                    style="thick",
                    color=COR_BORDA_SEQUENCIA,
                )
            )

    worksheet.freeze_panes = "A2"
    arquivo = BytesIO()
    workbook.save(arquivo)

    return arquivo.getvalue()


def gerar_planilha(
    tipo: str,
    value: dict,
    formato: str,
) -> tuple[bytes, str, str]:
    if formato == "json":
        return (
            json.dumps(value, ensure_ascii=False, indent=2).encode(),
            "application/json",
            "transcricao.json",
        )

    if formato == "csv":
        return (
            gerar_csv(tipo, value),
            "text/csv; charset=utf-8",
            "transcricao.csv",
        )

    if formato == "xlsx":
        return (
            gerar_xlsx(tipo, value),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "transcricao.xlsx",
        )

    raise ValueError("Formato de planilha não suportado.")
