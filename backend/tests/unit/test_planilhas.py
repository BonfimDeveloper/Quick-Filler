from io import BytesIO

from openpyxl import load_workbook

from app.models.cartao_ponto import CartaoPonto
from app.models.holerite import Holerite
from app.services.planilhas import (
    avisos_cartao,
    avisos_holerite,
    gerar_csv,
    gerar_xlsx,
)


def test_csv_cartao_cria_pares_conforme_maior_numero_de_batidas():
    value = {
        "pages": [
            {
                "page": 1,
                "days": [
                    {
                        "date_raw": "01/01/2020",
                        "punches": [
                            {"kind": "IN", "time_raw": "08:00", "time_hhmm": "08:00"},
                            {"kind": "OUT", "time_raw": "12:00", "time_hhmm": "12:00"},
                            {"kind": "IN", "time_raw": "13:00", "time_hhmm": "13:00"},
                        ],
                    }
                ],
            }
        ]
    }

    texto = gerar_csv("cartao-ponto", value).decode("utf-8-sig")

    assert "Data,Entrada 1,Saída 1,Entrada 2" in texto
    assert "01/01/2020,08:00,12:00,13:00" in texto


def test_xlsx_holerite_transpoe_verbas_e_formata_cabecalho():
    value = {
        "pages": [
            {
                "page": 1,
                "month": "01",
                "year": "2020",
                "fields": [
                    {"code": "001", "label": "Salário", "reference": "", "value": "1.000,00"}
                ],
                "bases": [],
            }
        ]
    }

    arquivo = gerar_xlsx("holerite", value)
    worksheet = load_workbook(BytesIO(arquivo)).active

    assert [cell.value for cell in worksheet[1]] == [
        "Pág.", "Mês", "Ano", "Salário"
    ]
    assert worksheet.cell(2, 4).value == "1.000,00"
    assert worksheet.cell(1, 1).fill.fgColor.rgb == "00173772"
    assert worksheet.cell(1, 1).alignment.wrap_text is True
    assert worksheet.row_dimensions[1].height == 42
    assert worksheet.column_dimensions["D"].width >= 10
    assert worksheet.auto_filter.ref == "A1:D2"


def test_avisos_cartao_priorizam_data_nao_sequencial():
    cartao = CartaoPonto.model_validate(
        {
            "pages": [
                {
                    "page": 1,
                    "days": [
                        {"date_raw": "01/01/2020", "punches": []},
                        {
                            "date_raw": "03/01/2020",
                            "punches": [
                                {"kind": "IN", "time_raw": "0?:00", "time_hhmm": "0?:00"}
                            ],
                        },
                    ],
                }
            ]
        }
    )

    assert avisos_cartao(cartao) == [None, "sequencia"]


def test_avisos_holerite_detectam_pagina_vazia_e_mes_pulado():
    holerite = Holerite.model_validate(
        {
            "pages": [
                {"page": 1, "month": "12", "year": "2019", "fields": [], "bases": []},
                {"page": 2, "month": "02", "year": "2020", "fields": [], "bases": []},
            ]
        }
    )

    assert avisos_holerite(holerite) == ["aviso", "sequencia"]


def test_xlsx_aplica_vermelho_e_borda_na_quebra_de_sequencia():
    value = {
        "pages": [
            {"page": 1, "month": "01", "year": "2020", "fields": [], "bases": []},
            {"page": 2, "month": "03", "year": "2020", "fields": [], "bases": []},
        ]
    }

    worksheet = load_workbook(
        BytesIO(gerar_xlsx("holerite", value))
    ).active

    assert worksheet.cell(3, 1).fill.fgColor.rgb == "00F8D7DA"
    assert worksheet.cell(3, 1).border.left.color.rgb == "00DC3545"
